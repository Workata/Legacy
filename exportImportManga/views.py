import io
from django.shortcuts import render
from django.shortcuts import redirect
from django.http import HttpResponse
from mangaDiary.models import MangaPersonal
from mangaDiary.models import MangaInfoPersonal
from mangaScraper.models import MangaGlobal
import xlsxwriter as excelWriter
import openpyxl

from mangaDiary.views import viewMangasList

# Create your views here.

#format2 = workbook.add_format({'num_format': 'dd/mm/yy'})
#worksheet.write('A2', number, format2)       # 28/02/13

def exportManga(request):

    user = request.user
    if not user.is_authenticated:
        return redirect('login')

    mangas = MangaPersonal.objects.filter(userId=user.id)

    # from stack
    output = io.BytesIO()
    workbook = excelWriter.Workbook(output, {'in_memory': True})
    dateFormat = workbook.add_format({'num_format': 'dd/mm/yy'})    # * SEE  https://xlsxwriter.readthedocs.io/working_with_dates_and_time.html
    worksheet = workbook.add_worksheet("Manga")                     # write (y,x)

    worksheet.write(0, 0, 'Title')      
    worksheet.write(0, 1, 'Volumes')  
    worksheet.write(0, 2, 'Status')  
    worksheet.write(0, 3, 'End date')  
    worksheet.write(0, 4, 'Rating')  
    worksheet.write(0, 5, 'Comment')   
    worksheet.write(0, 6, 'Full title')
    worksheet.write(0, 7, 'Written by')
    worksheet.write(0, 8, 'Illustrated by')
    worksheet.write(0, 9, 'Published by')
    worksheet.write(0, 10, 'English publisher')
    worksheet.write(0, 11, 'Imprint')
    worksheet.write(0, 12, 'Magazine')
    worksheet.write(0, 13, 'Demographic')
    worksheet.write(0, 14, 'Original run')   
    y = 1

    for manga in mangas:

        worksheet.write(y, 0, manga.title)
        worksheet.write(y, 1, str(manga.finishedVolumes) +"/"+ "0")  
        worksheet.write(y, 2, manga.status)
        worksheet.write(y, 3, manga.endDate, dateFormat)
        worksheet.write(y, 4, str(manga.rating))
        worksheet.write(y, 5, manga.comment)

        # * DESCRIPTION check if informations about this manga exist in database
        if manga.mangaInfoPersonalId != None:
            mangaInfo = MangaInfoPersonal.objects.filter(id=manga.mangaInfoPersonalId).first()
        elif manga.mangaGlobalId != None:
            mangaInfo = MangaGlobal.objects.filter(id=manga.mangaGlobalId).first()
        else:
            y += 1
            continue
        
        worksheet.write(y, 1, str(manga.finishedVolumes) +"/"+ str(mangaInfo.volumes))  
        worksheet.write(y, 6, mangaInfo.title)
        worksheet.write(y, 7, mangaInfo.writtenBy)  
        worksheet.write(y, 8, mangaInfo.illustratedBy)
        worksheet.write(y, 9, mangaInfo.publishedBy)
        worksheet.write(y, 10, mangaInfo.englishPublisher)
        worksheet.write(y, 11, mangaInfo.imprint)
        worksheet.write(y, 12, mangaInfo.magazine)
        worksheet.write(y, 13, mangaInfo.demographic)
        worksheet.write(y, 14, mangaInfo.originalRun)

        y += 1

    workbook.close()

    output.seek(0)

    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=mangaExport.xlsx"

    output.close()

    return response


# ??? Do I want overwrite data (using imported one) or no 
# * CAUTION right now its added to existing data 
def importManga(request):

    user = request.user
    if not user.is_authenticated:
        return redirect('login')

    if request.method == 'POST':                         # ??? why post lol
        excelFile = request.FILES['myfile']              # TODO validation if its really excel file
        workbook = openpyxl.load_workbook(excelFile)
        worksheet = workbook["Manga"]

        excelData = list()

        # iterating over the rows and getting value from each cell in row
        for row in worksheet.iter_rows():
            rowData = list()
            for cell in row:
                rowData.append(str(cell.value))
            excelData.append(rowData)

        excelData.pop(0)    # Delete row with column titles (first row)

        for row in excelData:

            manga = MangaPersonal()
            volumesString = row[1]

            # * DESCRIPTION check if manga contains additional informations (if yes create object from MangaInfoPersonal class)
            if row[6] != 'None' or row[7] != 'None' or row[8] != 'None' or row[9] != 'None' or row[10] != 'None' or row[11] != 'None' or row[12] != 'None' or row[13] != 'None' or row[14] != 'None' or volumesString[2] != '0':

                mangaInfo = MangaInfoPersonal()
                mangaInfo.title            = row[6]
                mangaInfo.writtenBy        = row[7]
                mangaInfo.illustratedBy    = row[8]
                mangaInfo.publishedBy      = row[9]
                mangaInfo.englishPublisher = row[10]
                mangaInfo.imprint          = row[11]
                mangaInfo.magazine         = row[12]
                mangaInfo.demographic      = row[13]
                mangaInfo.originalRun      = row[14]
                mangaInfo.volumes          = int(volumesString[2:])
                mangaInfo.save()

                manga.mangaInfoPersonalId = mangaInfo.id

            manga.title            = row[0]
            manga.finishedVolumes  = int(volumesString[0])   
            manga.status           = row[2]  
            endDateString          = row[3]             # 2020-08-24 00:00:00 (have to change this to -> 2020-08-24)
            endDate                = endDateString[:10]
            manga.endDate          = endDate            
            manga.rating           = int(row[4])
            manga.comment          = row[5]
            manga.userId           = user.id
            manga.save()
        
        return render(request, 'importManga.html')
    else:
        return render(request, 'importManga.html')