import io
from django.shortcuts import render
from django.shortcuts import redirect
from django.http import HttpResponse
from Diary.models import AnimePersonal
from Diary.models import AnimeInfoPersonal
from WikiScraper.models import AnimeGlobal
import xlsxwriter as excelWriter
import openpyxl

from Diary.views import viewAnimeList

# Create your views here.

#format2 = workbook.add_format({'num_format': 'dd/mm/yy'})
#worksheet.write('A2', number, format2)       # 28/02/13

def exportAnime(request):

    user = request.user
    if not user.is_authenticated:
        return redirect('login')

    animes = AnimePersonal.objects.filter(userId=user.id)

    output = io.BytesIO()

    workbook = excelWriter.Workbook(output, {'in_memory': True})
    dateFormat = workbook.add_format({'num_format': 'dd/mm/yy'})    # https://xlsxwriter.readthedocs.io/working_with_dates_and_time.html
    worksheet = workbook.add_worksheet("Anime") # write (y,x)

    worksheet.write(0, 0, 'Title')      
    worksheet.write(0, 1, 'Episodes')  
    worksheet.write(0, 2, 'Status')  
    worksheet.write(0, 3, 'End date')  
    worksheet.write(0, 4, 'Rating')  
    worksheet.write(0, 5, 'Comment')   
    worksheet.write(0, 6, 'Full title')
    worksheet.write(0, 7, 'Directed by')
    worksheet.write(0, 8, 'Produced by')
    worksheet.write(0, 9, 'Written by')
    worksheet.write(0, 10, 'Music by')
    worksheet.write(0, 11, 'Studio')
    worksheet.write(0, 12, 'Licensed by')
    worksheet.write(0, 13, 'Original network')
    worksheet.write(0, 14, 'Original run')   
    y = 1

    for anime in animes:

        worksheet.write(y, 0, anime.title)
        worksheet.write(y, 1, str(anime.finishedEpisodes) +"/"+ "0")  
        worksheet.write(y, 2, anime.status)
        worksheet.write(y, 3, anime.endDate, dateFormat)
        worksheet.write(y, 4, str(anime.rating))
        worksheet.write(y, 5, anime.status)

        print("AnimeInfoPersonalId:", anime.animeInfoPersonalId)
        print("AnimeInfoGlobalId:", anime.animeGlobalId)   

        if anime.animeInfoPersonalId != None:
            animeInfo = AnimeInfoPersonal.objects.filter(id=anime.animeInfoPersonalId).first()
        elif anime.animeGlobalId != None:
            animeInfo = AnimeGlobal.objects.filter(id=anime.animeGlobalId).first()
        else:
            y += 1
            continue
        
        worksheet.write(y, 1, str(anime.finishedEpisodes) +"/"+ str(animeInfo.episodes))  
        worksheet.write(y, 6, animeInfo.title)
        worksheet.write(y, 7, animeInfo.directedBy)  
        worksheet.write(y, 8, animeInfo.producedBy)
        worksheet.write(y, 9, animeInfo.writtenBy)
        worksheet.write(y, 10, animeInfo.musicBy)
        worksheet.write(y, 11, animeInfo.studio)
        worksheet.write(y, 12, animeInfo.licensedBy)
        worksheet.write(y, 13, animeInfo.originalNetwork)
        worksheet.write(y, 14, animeInfo.originalRun)

        y += 1

    workbook.close()

    output.seek(0)

    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=animeExport.xlsx"

    output.close()

    return response


def importAnime(request):

    user = request.user
    if not user.is_authenticated:
        return redirect('login')

    if request.method == 'POST':
        excelFile = request.FILES['myfile']       # TODO validation if its really excel file
        workbook = openpyxl.load_workbook(excelFile)
        worksheet = workbook["Anime"]

        excelData = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            rowData = list()
            for cell in row:
                rowData.append(str(cell.value))
            excelData.append(rowData)

        excelData.pop(0)    # Delete row with column titles (first row)
        for row in excelData:
            anime = AnimePersonal()
            anime.title = row[0]
            episodesString = row[1]
            anime.finishedEpisodes = int(episodesString[0])    #print(episodesStri)
            anime.status = row[2]  
            endDateString = row[3] #2020-08-24 00:00:00 (have to change this to -> 2020-08-24)
            endDate =  endDateString[:10]
            anime.endDate = endDate             #test
            anime.rating = int(row[4])
            anime.comment =  row[5]
            anime.userId = user.id
            anime.save()


        print(excelData[0])
        
        return render(request, 'importAnime.html')
    else:
        return render(request, 'importAnime.html')